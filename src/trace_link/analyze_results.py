import os
import sys
import json
import re
from collections import defaultdict
from typing import Dict, List
import glob

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '../../')))
from src.utils.utils import load_config

CONFIG = load_config()


class TraceLinkResultAnalyzer:
    def __init__(self, repo_name: str = None):
        self.repo_name = repo_name or CONFIG.get('repo', '')
        self.statistics_dir = os.path.join('data', self.repo_name, 'statistics_results')
        self.statistics_files = []
        self.all_results = []
        self.code_snippet_stats = {}
        self.overall_stats = {}

    def load_overall_stats(self) -> Dict:
        import glob as glob_module

        prompt_dir = os.path.join('data', '..', 'LLMapi', 'prompt', 'process_req')
        prompt_files = glob_module.glob(os.path.join(prompt_dir, "*.json"))

        overall = {
            'total_requirements': 0,
            'requirements_with_change_files': 0,
            'total_change_files': 0,
            'total_hit_files': 0,
            'avg_recall': 0.0,
            'prompt_count': 0
        }

        recalls = []
        for pf in prompt_files:
            try:
                with open(pf, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                stats = data.get('statistics', {})
                if stats:
                    overall['total_requirements'] = max(overall['total_requirements'], stats.get('total_requirements', 0))
                    overall['requirements_with_change_files'] = max(overall['requirements_with_change_files'], stats.get('requirements_with_change_files', 0))
                    overall['total_change_files'] = max(overall['total_change_files'], stats.get('total_change_files', 0))
                    overall['total_hit_files'] = max(overall['total_hit_files'], stats.get('total_hit_files', 0))
                    recalls.append(stats.get('overall_recall', 0))
                    overall['prompt_count'] += 1
            except Exception as e:
                pass

        if recalls:
            overall['avg_recall'] = sum(recalls) / len(recalls)

        self.overall_stats = overall
        return overall

    def load_code_snippet_stats(self) -> Dict:
        import torch
        sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '../../')))
        from src.model.calculate_code_vectors import get_pt_file_name

        pt_file_name = get_pt_file_name()
        pt_file_path = os.path.join('data', self.repo_name, pt_file_name)

        snippet_counts = {
            'total_snippets': 0,
            'CO': 0,
            'MO': 0,
            'IMO': 0,
            'MCC': 0,
            'MDCC': 0,
            'MC': 0,
            'IMD': 0,
            'IO': 0,
            'FC': 0,
            'CD': 0,
            'MD': 0,
            'IMC': 0,
            'other': 0
        }

        try:
            data = torch.load(pt_file_path, weights_only=False)
            snippet_types = data.get('snippet_types', [])

            snippet_counts['total_snippets'] += len(snippet_types)

            for st in snippet_types:
                if st.endswith('_CO') or st == 'class_CO':
                    snippet_counts['CO'] += 1
                elif st.endswith('_MO') or st == 'class_MO':
                    snippet_counts['MO'] += 1
                elif st == 'method_IMO':
                    snippet_counts['IMO'] += 1
                elif st == 'method_MCC':
                    snippet_counts['MCC'] += 1
                elif st == 'method_MDCC':
                    snippet_counts['MDCC'] += 1
                elif st == 'method_MC':
                    snippet_counts['MC'] += 1
                elif st == 'method_IMD':
                    snippet_counts['IMD'] += 1
                elif st == 'interface_IO' or st == 'class_IO':
                    snippet_counts['IO'] += 1
                elif st == 'code_FC':
                    snippet_counts['FC'] += 1
                elif st == 'class_CD':
                    snippet_counts['CD'] += 1
                elif st == 'method_MD':
                    snippet_counts['MD'] += 1
                elif st == 'method_IMC':
                    snippet_counts['IMC'] += 1
                else:
                    snippet_counts['other'] += 1
        except Exception as e:
            print(f"Error loading {pt_file_path}: {e}")

        self.code_snippet_stats = snippet_counts
        return snippet_counts

    def load_all_results(self) -> List[Dict]:
        self.statistics_files = glob.glob(os.path.join(self.statistics_dir, "*.json"))
        self.statistics_files = [f for f in self.statistics_files if not f.endswith('_analysis_output.json')]

        self.load_code_snippet_stats()
        self.load_overall_stats()

        self.all_results = []

        for filepath in self.statistics_files:
            filename = os.path.basename(filepath)
            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                config = self.parse_statistics_file(data, filename)
                config['source'] = 'statistics'
                config['filepath'] = filepath
                self.all_results.append({
                    'filename': filename,
                    'config': config,
                    'data': data,
                    'type': 'stats_only'
                })
            except Exception as e:
                print(f"Error loading {filepath}: {e}")

        return self.all_results

    def parse_trace_link_filename(self, filename: str) -> Dict:
        pro_match = re.match(r"trace_linkPro_(.+?)_(.+?)_(.+?)_top(\d+)_(\d+)(?:_(.+?))?\.json", filename)
        llm_no_pro_match = re.match(r"trace_link([a-zA-Z0-9_-]+?)_(jina_code|unixcoder|sbert|fastText)_top(\d+)_(\d+)(?:_(.+?))?\.json", filename)
        normal_match = re.match(r"trace_link([a-zA-Z0-9_-]+?)_top(\d+)_(\d+)(?:_(.+?))?\.json", filename)

        if pro_match:
            llm_provider = pro_match.group(1)
            model_name = pro_match.group(2)
            encoder = pro_match.group(3)
            top_k_min = int(pro_match.group(4))
            top_k_max = int(pro_match.group(5))
            snippet_types = pro_match.group(6).split("_") if pro_match.group(6) else []
            llm_name = "Pro_" + llm_provider + "_" + model_name.replace("_", "/")
            return {
                'llm_name': llm_name,
                'encoder': encoder,
                'top_k_range': f"{top_k_min}-{top_k_max}",
                'top_k_min': top_k_min,
                'top_k_max': top_k_max,
                'snippet_types': snippet_types,
                'snippet_types_str': "_".join(snippet_types) if snippet_types else "default",
                'has_llm': True,
                'use_llm': llm_name
            }
        elif llm_no_pro_match:
            llm_name = llm_no_pro_match.group(1)
            encoder = llm_no_pro_match.group(2)
            top_k_min = int(llm_no_pro_match.group(3))
            top_k_max = int(llm_no_pro_match.group(4))
            snippet_types = llm_no_pro_match.group(5).split("_") if llm_no_pro_match.group(5) else []
            return {
                'llm_name': llm_name,
                'encoder': encoder,
                'top_k_range': f"{top_k_min}-{top_k_max}",
                'top_k_min': top_k_min,
                'top_k_max': top_k_max,
                'snippet_types': snippet_types,
                'snippet_types_str': "_".join(snippet_types) if snippet_types else "default",
                'has_llm': True,
                'use_llm': llm_name
            }
        elif normal_match:
            encoder = normal_match.group(1)
            top_k_min = int(normal_match.group(2))
            top_k_max = int(normal_match.group(3))
            snippet_types = normal_match.group(4).split("_") if normal_match.group(4) else []
            return {
                'llm_name': None,
                'encoder': encoder,
                'top_k_range': f"{top_k_min}-{top_k_max}",
                'top_k_min': top_k_min,
                'top_k_max': top_k_max,
                'snippet_types': snippet_types,
                'snippet_types_str': "_".join(snippet_types) if snippet_types else "default",
                'has_llm': False,
                'use_llm': "未使用LLM"
            }
        else:
            return {
                'llm_name': None,
                'encoder': 'unknown',
                'top_k_range': 'unknown',
                'top_k_min': 0,
                'top_k_max': 0,
                'snippet_types': [],
                'snippet_types_str': 'unknown',
                'has_llm': False,
                'use_llm': "未使用LLM"
            }

    def parse_statistics_file(self, data: Dict, filename: str) -> Dict:
        config = data.get('config', {})
        req_proc = config.get('requirement_processing', {})
        trace_cfg = config.get('trace_link', {})
        use_llm = req_proc.get('use_llm_processing', False)

        llm_provider = config.get('LLMProvider')
        llm_model = None
        if llm_provider:
            llm_provider_config = config.get(llm_provider, {})
            llm_model = llm_provider_config.get('model')

        if not llm_model:
            prompt_name = req_proc.get('prompt_name', '')
            if prompt_name:
                llm_model = f"LLM({prompt_name})"

        return {
            'llm_name': llm_model,
            'encoder': config.get('encode_model_name', 'unknown'),
            'code_snippet': config.get('code_snippet', []),
            'snippet_types_str': "_".join(config.get('code_snippet', [])) if config.get('code_snippet') else "default",
            'top_k': config.get('top_k', []),
            'top_k_range': f"{min(config.get('top_k', [0]))}-{max(config.get('top_k', [0]))}" if config.get('top_k') else "unknown",
            'top_k_min': min(config.get('top_k', [0])) if config.get('top_k') else 0,
            'top_k_max': max(config.get('top_k', [0])) if config.get('top_k') else 0,
            'unique_file_only': config.get('unique_file_only', True),
            'use_llm_processing': use_llm,
            'trace_link_use_llm': trace_cfg.get('use_llm', False),
            'has_llm': use_llm or trace_cfg.get('use_llm', False),
            'use_llm': llm_model if (use_llm or trace_cfg.get('use_llm', False)) else "未使用LLM",
            'repo': config.get('repo', 'unknown'),
            'requirement_processing': req_proc,
            'trace_link': trace_cfg,
            'LLMProvider': llm_provider,
            'LLMtemperature': config.get('LLMtemperature'),
            'code_embedding': config.get('code_embedding', {}),
            'prompt_name': req_proc.get('prompt_name', '') if use_llm else '',
            'prefix_title': 'YES' if req_proc.get('prefix_title', False) else 'NO'
        }

    def get_all_statistics(self) -> List[Dict]:
        rows = []
        for result in self.all_results:
            config = result['config']
            data = result['data']

            if result['type'] == 'stats_only':
                statistics = data.get('statistics', {})
                for top_k, stats in statistics.items():
                    hit_rate = stats.get('requirements_with_at_least_one_hit', 0) / stats.get('requirements_with_change_files', 1) if stats.get('requirements_with_change_files', 0) > 0 else 0.0
                    rows.append({
                        'filename': result['filename'],
                        'source': 'statistics',
                        'llm_name': config.get('llm_name'),
                        'encoder': config.get('encoder', 'unknown'),
                        'snippet_types_str': config.get('snippet_types_str', 'unknown'),
                        'prompt_name': config.get('prompt_name', ''),
                        'prefix_title': config.get('prefix_title', ''),
                        'LLMtemperature': config.get('LLMtemperature', ''),
                        'top_k': int(top_k),
                        'has_llm': config.get('has_llm', False),
                        'use_llm': config.get('use_llm', '未使用LLM'),
                        'total_requirements': stats.get('total_requirements', 0),
                        'requirements_with_change_files': stats.get('requirements_with_change_files', 0),
                        'requirements_with_at_least_one_hit': stats.get('requirements_with_at_least_one_hit', 0),
                        'total_change_files': stats.get('total_change_files', 0),
                        'total_predicted_files': stats.get('total_predicted_files', 0),
                        'total_hit_files': stats.get('total_hit_files', 0),
                        'total_fp_files': stats.get('total_fp_files', 0),
                        'overall_recall': stats.get('overall_recall', 0.0),
                        'overall_precision': stats.get('overall_precision', 0.0),
                        'overall_f1': stats.get('overall_f1', 0.0),
                        'average_recall': stats.get('average_recall', 0.0),
                        'average_precision': stats.get('average_precision', 0.0),
                        'average_f1': stats.get('average_f1', 0.0),
                        'hit_rate': hit_rate
                    })
            else:
                statistics = data.get('statistics', {})
                for top_k, stats in statistics.items():
                    hit_rate = stats.get('requirements_with_at_least_one_hit', 0) / stats.get('requirements_with_change_files', 1) if stats.get('requirements_with_change_files', 0) > 0 else 0.0
                    rows.append({
                        'filename': result['filename'],
                        'source': 'trace_link',
                        'llm_name': config.get('llm_name'),
                        'encoder': config.get('encoder', 'unknown'),
                        'snippet_types_str': config.get('snippet_types_str', 'unknown'),
                        'top_k': int(top_k),
                        'has_llm': config.get('has_llm', False),
                        'use_llm': config.get('use_llm', '未使用LLM'),
                        'total_requirements': stats.get('total_requirements', 0),
                        'requirements_with_change_files': stats.get('requirements_with_change_files', 0),
                        'requirements_with_at_least_one_hit': stats.get('requirements_with_at_least_one_hit', 0),
                        'total_change_files': stats.get('total_change_files', 0),
                        'total_predicted_files': stats.get('total_predicted_files', 0),
                        'total_hit_files': stats.get('total_hit_files', 0),
                        'total_fp_files': stats.get('total_fp_files', 0),
                        'overall_recall': stats.get('overall_recall', 0.0),
                        'overall_precision': stats.get('overall_precision', 0.0),
                        'overall_f1': stats.get('overall_f1', 0.0),
                        'average_recall': stats.get('average_recall', 0.0),
                        'average_precision': stats.get('average_precision', 0.0),
                        'average_f1': stats.get('average_f1', 0.0),
                        'hit_rate': hit_rate
                    })

        return rows

    def compare_top_k_performance(self) -> List[Dict]:
        stats = self.get_all_statistics()
        topk_stats = defaultdict(list)
        for s in stats:
            topk_stats[s['top_k']].append(s)

        results = []
        for top_k, stat_list in sorted(topk_stats.items()):
            avg_recall = sum(s['overall_recall'] for s in stat_list) / len(stat_list)
            avg_precision = sum(s['overall_precision'] for s in stat_list) / len(stat_list)
            avg_f1 = sum(s['overall_f1'] for s in stat_list) / len(stat_list)
            avg_avg_recall = sum(s['average_recall'] for s in stat_list) / len(stat_list)
            avg_avg_precision = sum(s['average_precision'] for s in stat_list) / len(stat_list)
            avg_avg_f1 = sum(s['average_f1'] for s in stat_list) / len(stat_list)
            avg_hit_rate = sum(s['hit_rate'] for s in stat_list) / len(stat_list)
            results.append({
                'top_k': top_k,
                'overall_recall': round(avg_recall, 4),
                'overall_precision': round(avg_precision, 4),
                'overall_f1': round(avg_f1, 4),
                'average_recall': round(avg_avg_recall, 4),
                'average_precision': round(avg_avg_precision, 4),
                'average_f1': round(avg_avg_f1, 4),
                'hit_rate': round(avg_hit_rate, 4),
                'count': len(stat_list)
            })
        return results

    def compare_encoders(self, top_k: int = 10) -> List[Dict]:
        stats = self.get_all_statistics()
        filtered = [s for s in stats if s['top_k'] == top_k]

        encoder_stats = defaultdict(list)
        for s in filtered:
            key = (s['encoder'], s['snippet_types_str'], s['use_llm'])
            encoder_stats[key].append(s)

        results = []
        for (encoder, snippet, use_llm), stat_list in encoder_stats.items():
            avg_recall = sum(s['overall_recall'] for s in stat_list) / len(stat_list)
            avg_precision = sum(s['overall_precision'] for s in stat_list) / len(stat_list)
            avg_f1 = sum(s['overall_f1'] for s in stat_list) / len(stat_list)
            avg_avg_recall = sum(s['average_recall'] for s in stat_list) / len(stat_list)
            avg_avg_precision = sum(s['average_precision'] for s in stat_list) / len(stat_list)
            avg_avg_f1 = sum(s['average_f1'] for s in stat_list) / len(stat_list)
            avg_hit_rate = sum(s['hit_rate'] for s in stat_list) / len(stat_list)
            results.append({
                'encoder': encoder,
                'snippet_types': snippet,
                'use_llm': use_llm,
                'overall_recall': round(avg_recall, 4),
                'overall_precision': round(avg_precision, 4),
                'overall_f1': round(avg_f1, 4),
                'average_recall': round(avg_avg_recall, 4),
                'average_precision': round(avg_avg_precision, 4),
                'average_f1': round(avg_avg_f1, 4),
                'hit_rate': round(avg_hit_rate, 4),
                'count': len(stat_list)
            })

        return sorted(results, key=lambda x: x['overall_recall'], reverse=True)

    def compare_llm_usage(self, top_k: int = 10) -> List[Dict]:
        stats = self.get_all_statistics()
        filtered = [s for s in stats if s['top_k'] == top_k]

        llm_stats = defaultdict(list)
        for s in filtered:
            llm_stats[s['use_llm']].append(s)

        results = []
        for use_llm, stat_list in llm_stats.items():
            avg_recall = sum(s['overall_recall'] for s in stat_list) / len(stat_list)
            avg_precision = sum(s['overall_precision'] for s in stat_list) / len(stat_list)
            avg_f1 = sum(s['overall_f1'] for s in stat_list) / len(stat_list)
            results.append({
                'use_llm': use_llm,
                'overall_recall': round(avg_recall, 4),
                'overall_precision': round(avg_precision, 4),
                'overall_f1': round(avg_f1, 4),
                'count': len(stat_list)
            })

        return sorted(results, key=lambda x: x['overall_recall'], reverse=True)

    def compare_prompt_performance(self, top_k: int = 10) -> List[Dict]:
        stats = self.get_all_statistics()
        filtered = [s for s in stats if s['top_k'] == top_k]

        prompt_stats = defaultdict(list)
        for s in filtered:
            prompt_name = s.get('prompt_name', '') or '未使用LLM'
            prefix_title = s.get('prefix_title', 'NO')
            key = f"{prompt_name}_{prefix_title}"
            prompt_stats[key].append(s)

        results = []
        for key, stat_list in prompt_stats.items():
            avg_recall = sum(s['overall_recall'] for s in stat_list) / len(stat_list)
            avg_precision = sum(s['overall_precision'] for s in stat_list) / len(stat_list)
            avg_f1 = sum(s['overall_f1'] for s in stat_list) / len(stat_list)
            parts = key.rsplit('_', 1)
            prompt_name = parts[0] if len(parts) > 0 else key
            prefix_title = parts[1] if len(parts) > 1 else 'NO'
            results.append({
                'prompt_name': prompt_name,
                'prefix_title': prefix_title,
                'overall_recall': round(avg_recall, 4),
                'overall_precision': round(avg_precision, 4),
                'overall_f1': round(avg_f1, 4),
                'count': len(stat_list)
            })

        return sorted(results, key=lambda x: x['overall_recall'], reverse=True)

    def print_code_snippet_stats(self):
        print("=" * 100)
        print("代码片段统计 (.pt文件)")
        print("=" * 100)
        stats = self.code_snippet_stats
        if not stats:
            print("未找到代码片段统计数据")
            return
        print(f"\n总片段数: {stats.get('total_snippets', 0):,}")
        print("\n各类片段数量:")
        print(f"  CO (类代码片段):     {stats.get('CO', 0):>8,}")
        print(f"  MO (方法代码片段):   {stats.get('MO', 0):>8,}")
        print(f"  IMO (接口方法片段):  {stats.get('IMO', 0):>8,}")
        print(f"  MCC (方法类上下文):  {stats.get('MCC', 0):>8,}")
        print(f"  MDCC (方法注释类上下文): {stats.get('MDCC', 0):>8,}")
        print(f"  MC (方法注释):       {stats.get('MC', 0):>8,}")
        print(f"  IMD (接口方法注释):  {stats.get('IMD', 0):>8,}")
        print(f"  IO (接口片段):       {stats.get('IO', 0):>8,}")
        print(f"  FC (文件代码片段):   {stats.get('FC', 0):>8,}")
        print(f"  CD (类注释片段):     {stats.get('CD', 0):>8,}")
        print(f"  MD (方法注释):       {stats.get('MD', 0):>8,}")
        print(f"  IMC (接口方法注释):  {stats.get('IMC', 0):>8,}")
        print(f"  other:               {stats.get('other', 0):>8,}")
        print()

    def print_overall_stats(self):
        print("=" * 100)
        print("整体项目统计")
        print("=" * 100)
        stats = self.overall_stats
        if not stats:
            print("未找到整体统计数据")
            return
        print(f"\n  总需求数:                    {stats.get('total_requirements', 0):>8,}")
        print(f"  有变更文件的需求数:          {stats.get('requirements_with_change_files', 0):>8,}")
        print(f"  总变更文件数:                {stats.get('total_change_files', 0):>8,}")
        print(f"  总命中文件数:                {stats.get('total_hit_files', 0):>8,}")
        print(f"  测试的Prompt数量:            {stats.get('prompt_count', 0):>8,}")
        print(f"  平均召回率 (from prompts):   {stats.get('avg_recall', 0):>8.4f}")
        print()

    def print_summary(self):
        print("=" * 100)
        print(f"需求追踪链接结果分析 - {self.repo_name}")
        print("=" * 100)
        print(f"\n共加载 {len(self.statistics_files)} 个统计文件\n")

        self.print_code_snippet_stats()
        self.print_overall_stats()

        all_stats = self.get_all_statistics()

        print("=" * 100)
        print("LLM使用对比 (Top-10, 按Recall排序)")
        print("=" * 100)
        llm_comparison = self.compare_llm_usage(top_k=10)
        header = f"{'UseLLM':>30} | {'Recall':>8} | {'Precision':>10} | {'F1':>8} | {'Count':>6}"
        print(header)
        print("-" * 70)
        for row in llm_comparison:
            print(f"{row['use_llm']:>30} | {row['overall_recall']:>8.4f} | {row['overall_precision']:>10.4f} | {row['overall_f1']:>8.4f} | {row['count']:>6}")

        print("\n" + "=" * 100)
        print("Prompt对比 (Top-10, 按Recall排序)")
        print("=" * 100)
        prompt_comparison = self.compare_prompt_performance(top_k=10)
        header = f"{'Prompt':>20} | {'PrefixTitle':>12} | {'Recall':>8} | {'Precision':>10} | {'F1':>8} | {'Count':>6}"
        print(header)
        print("-" * 75)
        for row in prompt_comparison:
            print(f"{row['prompt_name']:>20} | {row['prefix_title']:>12} | {row['overall_recall']:>8.4f} | {row['overall_precision']:>10.4f} | {row['overall_f1']:>8.4f} | {row['count']:>6}")

        print("\n" + "=" * 100)
        print("Top-K 性能对比 (各配置平均)")
        print("=" * 100)
        df_topk = self.compare_top_k_performance()
        header = f"{'Top-K':>6} | {'Recall':>8} | {'Precision':>10} | {'F1':>8} | {'Avg Recall':>10} | {'Avg Prec':>10} | {'Avg F1':>8} | {'Hit Rate':>9}"
        print(header)
        print("-" * 100)
        for row in df_topk:
            print(f"{row['top_k']:>6} | {row['overall_recall']:>8.4f} | {row['overall_precision']:>10.4f} | {row['overall_f1']:>8.4f} | {row['average_recall']:>10.4f} | {row['average_precision']:>10.4f} | {row['average_f1']:>8.4f} | {row['hit_rate']:>9.4f}")

        print("\n" + "=" * 100)
        print("编码器+LLM组合对比 (Top-10, 按Recall排序)")
        print("=" * 100)
        df_encoder = self.compare_encoders(top_k=10)
        header = f"{'Encoder':>15} | {'Snippet Types':>25} | {'UseLLM':>20} | {'Recall':>8} | {'F1':>8}"
        print(header)
        print("-" * 100)
        for row in df_encoder[:20]:
            print(f"{row['encoder']:>15} | {row['snippet_types']:>25} | {row['use_llm']:>20} | {row['overall_recall']:>8.4f} | {row['overall_f1']:>8.4f}")

        print("\n" + "=" * 100)
        print("各 Top-K 配置详细结果 (按Recall排序)")
        print("=" * 100)
        for top_k in sorted(set(s['top_k'] for s in all_stats)):
            df_k = [s for s in all_stats if s['top_k'] == top_k]
            df_k.sort(key=lambda x: x['overall_recall'], reverse=True)
            print(f"\n--- Top-{top_k} ---")
            header = f"{'Encoder':>12} | {'Snippet Types':>25} | {'UseLLM':>20} | {'Recall':>8} | {'Prec':>8} | {'F1':>8} | {'Count':>5}"
            print(header)
            print("-" * 100)
            for row in df_k:
                print(f"{row['encoder']:>12} | {row['snippet_types_str']:>25} | {row['use_llm']:>20} | {row['overall_recall']:>8.4f} | {row['overall_precision']:>8.4f} | {row['overall_f1']:>8.4f} | {row.get('count', 1):>5}")

    def export_to_json(self, output_path: str):
        all_stats = self.get_all_statistics()
        export_data = {
            'summary': {
                'repo': self.repo_name,
                'statistics_files': len(self.statistics_files),
                'code_snippet_stats': self.code_snippet_stats,
                'overall_stats': self.overall_stats
            },
            'llm_comparison': self.compare_llm_usage(top_k=10),
            'prompt_comparison': self.compare_prompt_performance(top_k=10),
            'top_k_comparison': self.compare_top_k_performance(),
            'encoder_comparison': self.compare_encoders(top_k=10),
            'all_statistics': all_stats
        }

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False)

        print(f"\nJSON结果已导出到: {output_path}")

    def export_to_excel(self, output_path: str):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            print("请先安装 openpyxl: pip install openpyxl")
            return

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font_white = Font(bold=True, color="FFFFFF")

        all_stats = self.get_all_statistics()

        ws_snippet = wb.create_sheet("代码片段统计")
        ws_snippet.append(["类型", "数量"])
        for col in range(1, 3):
            cell = ws_snippet.cell(row=1, column=col)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        ws_snippet.append(["总片段数", self.code_snippet_stats.get('total_snippets', 0)])
        ws_snippet.append(["CO (类代码片段)", self.code_snippet_stats.get('CO', 0)])
        ws_snippet.append(["MO (方法代码片段)", self.code_snippet_stats.get('MO', 0)])
        ws_snippet.append(["IMO (接口方法片段)", self.code_snippet_stats.get('IMO', 0)])
        ws_snippet.append(["MCC (方法类上下文)", self.code_snippet_stats.get('MCC', 0)])
        ws_snippet.append(["MDCC (方法注释类上下文)", self.code_snippet_stats.get('MDCC', 0)])
        ws_snippet.append(["MC (方法注释)", self.code_snippet_stats.get('MC', 0)])
        ws_snippet.append(["IMD (接口方法注释)", self.code_snippet_stats.get('IMD', 0)])
        ws_snippet.append(["IO (接口片段)", self.code_snippet_stats.get('IO', 0)])
        ws_snippet.append(["FC (文件代码片段)", self.code_snippet_stats.get('FC', 0)])
        ws_snippet.append(["CD (类注释片段)", self.code_snippet_stats.get('CD', 0)])
        ws_snippet.append(["MD (方法注释)", self.code_snippet_stats.get('MD', 0)])
        ws_snippet.append(["IMC (接口方法注释)", self.code_snippet_stats.get('IMC', 0)])
        ws_snippet.append(["other", self.code_snippet_stats.get('other', 0)])
        ws_snippet.column_dimensions['A'].width = 25
        ws_snippet.column_dimensions['B'].width = 15

        ws_overall = wb.create_sheet("整体情况")
        ws_overall.append(["指标", "值"])
        for col in range(1, 3):
            cell = ws_overall.cell(row=1, column=col)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        ws_overall.append(["项目", self.repo_name])
        ws_overall.append(["总需求数", self.overall_stats.get('total_requirements', 0)])
        ws_overall.append(["有变更文件的需求数", self.overall_stats.get('requirements_with_change_files', 0)])
        ws_overall.append(["总变更文件数", self.overall_stats.get('total_change_files', 0)])
        ws_overall.append(["总命中文件数", self.overall_stats.get('total_hit_files', 0)])
        ws_overall.append(["测试的Prompt数量", self.overall_stats.get('prompt_count', 0)])
        ws_overall.append(["平均召回率 (from prompts)", f"{self.overall_stats.get('avg_recall', 0):.4f}"])
        ws_overall.append(["代码片段总数", self.code_snippet_stats.get('total_snippets', 0)])
        ws_overall.column_dimensions['A'].width = 25
        ws_overall.column_dimensions['B'].width = 20

        ws_summary = wb.create_sheet("LLM对比")
        headers_summary = ["LLM类型", "Recall", "Precision", "F1", "数量"]
        ws_summary.append(headers_summary)
        for col in range(1, len(headers_summary) + 1):
            cell = ws_summary.cell(row=1, column=col)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for row in self.compare_llm_usage(top_k=10):
            ws_summary.append([
                row['use_llm'],
                f"{row['overall_recall']:.2%}",
                f"{row['overall_precision']:.2%}",
                f"{row['overall_f1']:.2%}",
                row['count']
            ])
        ws_summary.column_dimensions['A'].width = 30

        ws_prompt = wb.create_sheet("Prompt对比")
        headers_prompt = ["Prompt", "PrefixTitle", "Recall", "Precision", "F1", "数量"]
        ws_prompt.append(headers_prompt)
        for col in range(1, len(headers_prompt) + 1):
            cell = ws_prompt.cell(row=1, column=col)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for row in self.compare_prompt_performance(top_k=10):
            ws_prompt.append([
                row['prompt_name'],
                row['prefix_title'],
                f"{row['overall_recall']:.2%}",
                f"{row['overall_precision']:.2%}",
                f"{row['overall_f1']:.2%}",
                row['count']
            ])
        ws_prompt.column_dimensions['A'].width = 20
        ws_prompt.column_dimensions['B'].width = 12

        ws_topk = wb.create_sheet("Top-K汇总")
        headers_topk = ["Top-K", "Recall", "Precision", "F1", "Avg Recall", "Avg Prec", "Avg F1", "Hit Rate", "Count"]
        ws_topk.append(headers_topk)
        for col in range(1, len(headers_topk) + 1):
            cell = ws_topk.cell(row=1, column=col)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for row in self.compare_top_k_performance():
            ws_topk.append([
                row['top_k'],
                f"{row['overall_recall']:.2%}",
                f"{row['overall_precision']:.2%}",
                f"{row['overall_f1']:.2%}",
                f"{row['average_recall']:.2%}",
                f"{row['average_precision']:.2%}",
                f"{row['average_f1']:.2%}",
                f"{row['hit_rate']:.2%}",
                row['count']
            ])
        ws_topk.column_dimensions['A'].width = 8

        for top_k in sorted(set(s['top_k'] for s in all_stats)):
            df_k = [s for s in all_stats if s['top_k'] == top_k]
            df_k.sort(key=lambda x: x['overall_f1'], reverse=True)
            ws = wb.create_sheet(f"Top{top_k}详情")
            headers = ["Encoder", "Snippet Types", "UseLLM", "PromptName", "PrefixTitle", "Temperature", "Recall", "Precision", "F1", "Avg F1", "Hit Rate",
                      "总需求", "有变更", "至少命中", "变更文件", "预测文件", "命中", "FP"]
            ws.append(headers)
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            for row in df_k:
                ws.append([
                    row['encoder'],
                    row['snippet_types_str'],
                    row['use_llm'],
                    row.get('prompt_name', ''),
                    row.get('prefix_title', ''),
                    row.get('LLMtemperature', ''),
                    f"{row['overall_recall']:.2%}",
                    f"{row['overall_precision']:.2%}",
                    f"{row['overall_f1']:.2%}",
                    f"{row['average_f1']:.2%}",
                    f"{row['hit_rate']:.2%}",
                    row['total_requirements'],
                    row['requirements_with_change_files'],
                    row['requirements_with_at_least_one_hit'],
                    row['total_change_files'],
                    row['total_predicted_files'],
                    row['total_hit_files'],
                    row['total_fp_files']
                ])
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 25

        wb.save(output_path)
        print(f"\nExcel结果已导出到: {output_path}")


def main():
    repo_name = CONFIG['repo']
    data_dir = os.path.join('data')
    actual_dir = None
    if os.path.isdir(data_dir):
        for d in os.listdir(data_dir):
            if d.lower() == repo_name.lower():
                actual_dir = d
                break
    if actual_dir is None:
        actual_dir = repo_name
    results_dir = os.path.join('data', actual_dir, 'trace_link_results')
    stats_dir = os.path.join('data', actual_dir, 'statistics_results')

    os.makedirs(stats_dir, exist_ok=True)
    os.makedirs(results_dir, exist_ok=True)
    json_output = os.path.join(stats_dir, f"{actual_dir.lower()}_analysis_output.json")
    excel_output = os.path.join(results_dir, f"{actual_dir.lower()}_analysis_output.xlsx")

    analyzer = TraceLinkResultAnalyzer(repo_name)
    analyzer.load_all_results()
    analyzer.print_summary()
    analyzer.export_to_json(json_output)
    analyzer.export_to_excel(excel_output)


if __name__ == "__main__":
    main()
